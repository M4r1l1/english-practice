#!/usr/bin/env python3
"""
Script para sincronizar los términos del Excel con el archivo HTML del quiz.
Uso: python3 sync_excel_to_html.py

El Excel debe tener las columnas:
- Columna A: Término
- Columna B: Definición
- Columna C (opcional): Categoría (idiom, phrasal, vocab, collocation)

Si no hay categoría, el script intentará detectarla automáticamente.
"""

import re
from openpyxl import load_workbook
import os

def detect_category(term, definition):
    """Detecta automáticamente la categoría basada en el término."""
    term_lower = term.lower()
    def_lower = definition.lower()

    # Phrasal verbs - típicamente 2-3 palabras con verbo + preposición/adverbio
    phrasal_indicators = ['up', 'down', 'out', 'in', 'on', 'off', 'over', 'away', 'through', 'back', 'along']
    words = term_lower.split()
    if len(words) >= 2 and any(word in phrasal_indicators for word in words[1:]):
        if len(words) <= 4:  # Phrasal verbs are usually short
            return "phrasal"

    # Collocations - frases comunes con verbos específicos
    collocation_verbs = ['make', 'take', 'draw', 'reach', 'raise', 'pose', 'meet', 'bear', 'come', 'shed', 'pay', 'catch']
    if any(term_lower.startswith(verb + ' ') for verb in collocation_verbs):
        if len(words) >= 2 and len(words) <= 4:
            return "collocation"

    # Idioms - frases más largas con artículos o lenguaje figurativo
    idiom_indicators = ['the', 'a ', 'an ', 'your', 'one\'s']
    if len(words) >= 3 or any(ind in term_lower for ind in idiom_indicators):
        return "idiom"

    # Default to vocabulary
    return "vocab"

def read_excel_terms(excel_path):
    """Lee los términos del archivo Excel."""
    wb = load_workbook(excel_path)
    ws = wb.active

    terms = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        term = row[0]
        definition = row[1]
        category = row[2] if len(row) > 2 and row[2] else None

        if term and definition:
            term = str(term).strip()
            definition = str(definition).strip()
            example = row[3] if len(row) > 3 and row[3] else ""

            if not category:
                category = detect_category(term, definition)
            else:
                category = str(category).strip().lower()

            terms.append({
                'term': term,
                'definition': definition,
                'category': category,
                'example': str(example).strip() if example else ""
            })

    return terms

def generate_js_array(terms):
    """Genera el array JavaScript con los términos."""
    lines = ['        const allTerms = [']

    for i, t in enumerate(terms):
        # Escapar comillas en los strings
        term = t['term'].replace('"', '\\"').replace("'", "\\'")
        definition = t['definition'].replace('"', '\\"').replace("'", "\\'")
        category = t['category']
        example = t.get('example', '').replace('"', '\\"').replace("'", "\\'")

        comma = ',' if i < len(terms) - 1 else ''
        lines.append(f'            {{ term: "{term}", definition: "{definition}", category: "{category}", example: "{example}" }}{comma}')

    lines.append('        ];')
    return '\n'.join(lines)

def update_html_file(html_path, terms):
    """Actualiza el archivo HTML con los nuevos términos."""
    with open(html_path, 'r', encoding='utf-8') as f:
        content = f.read()

    # Patrón para encontrar el array de términos
    pattern = r'const allTerms = \[[\s\S]*?\];'

    new_array = generate_js_array(terms)

    new_content = re.sub(pattern, new_array, content)

    with open(html_path, 'w', encoding='utf-8') as f:
        f.write(new_content)

    return len(terms)

def main():
    # Rutas de archivos
    script_dir = os.path.dirname(os.path.abspath(__file__))
    excel_path = os.path.join(script_dir, 'excel', 'EnglishTerms.xlsx')
    html_path = os.path.join(script_dir, 'index.html')

    print("=" * 50)
    print("Sincronizador Excel -> HTML Quiz")
    print("=" * 50)

    # Verificar que existen los archivos
    if not os.path.exists(excel_path):
        print(f"Error: No se encontró el archivo Excel en {excel_path}")
        return

    if not os.path.exists(html_path):
        print(f"Error: No se encontró el archivo HTML en {html_path}")
        return

    print(f"\nLeyendo términos de: {excel_path}")
    terms = read_excel_terms(excel_path)

    # Mostrar estadísticas por categoría
    categories = {}
    for t in terms:
        cat = t['category']
        categories[cat] = categories.get(cat, 0) + 1

    print(f"\nTérminos encontrados: {len(terms)}")
    print("\nPor categoría:")
    for cat, count in sorted(categories.items()):
        print(f"  - {cat}: {count}")

    print(f"\nActualizando: {html_path}")
    count = update_html_file(html_path, terms)

    print(f"\n✓ Sincronización completada!")
    print(f"  {count} términos actualizados en el quiz.")
    print("=" * 50)

if __name__ == '__main__':
    main()
